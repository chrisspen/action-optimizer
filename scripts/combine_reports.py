#!/usr/bin/env python
import os
import sys
from datetime import date
from collections import OrderedDict
import pathlib

from pyexcel_ods import save_data
from pandas_ods_reader import read_ods
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

# pylint: disable=wrong-import-position

sys.path.insert(1, '..')
from action_optimizer.optimizer import Optimizer

current_path = pathlib.Path(__file__).parent.resolve()

dt = date.today()
if len(sys.argv) > 1:
    dt = sys.argv[1]

MAX_HEADER_SIZE = {
    'rec_value': 3,
    'rec_change': 1,
    'human_conf': 0,
}


def auto_size_and_fix_columns(input_ods, postfix=None):
    # Generate the output_xlsx filename
    postfix = postfix or ''
    output_xlsx = os.path.splitext(input_ods)[0] + postfix + '.xlsx'

    # Load the existing ODS file as a DataFrame
    df = pd.read_excel(input_ods, engine='odf')

    # Save the DataFrame to an Excel file (since openpyxl works with Excel files)
    df.to_excel(output_xlsx, index=False)

    # Load the Excel file with openpyxl
    wb = load_workbook(output_xlsx)
    ws = wb.active

    # Freeze the top header row
    ws.freeze_panes = ws['B2']

    # Apply bold formatting to the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Define a boolean style
    bool_style = NamedStyle(name="bool_style", number_format='BOOLEAN')
    if "bool_style" not in wb.named_styles:
        wb.add_named_style(bool_style)
    rec_change_col = None
    for cell in ws[1]:
        if cell.value == "rec_change":
            rec_change_col = cell.column_letter
            break

    # Auto-size the columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for i, cell in enumerate(col):
            if not i and cell.value.strip() in MAX_HEADER_SIZE:
                max_length = MAX_HEADER_SIZE[cell.value.strip()] + len(str(cell.value))
                break
            max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 1) # buffer
        ws.column_dimensions[column].width = adjusted_width

    # Apply the boolean style to the "rec_change" column
    if rec_change_col:
        for cell in ws[rec_change_col]:
            if cell.row != 1: # Skip the header
                cell.style = bool_style

    wb.save(output_xlsx)


def combine():

    datafile = os.environ.get('ACTION_OPTIMIZER_DATAFILE')
    print(f'Loading data file: {datafile}')
    optimizer = Optimizer(datafile)
    print('Getting first non-blank row.')
    last_data = optimizer.get_first_nonblank_row(headers=True)
    print('last_data:', last_data)
    assert last_data, 'Unable to find last non-blank row?!'

    pcc_path = os.path.abspath(f'{current_path}/../reports/{dt}/pcc.ods')
    print(f'Using PCC path: {pcc_path}')
    assert os.path.isfile(pcc_path)

    ens_path = os.path.abspath(f'{current_path}/../reports/{dt}/ensemble.ods')
    print(f'Using ENS path: {ens_path}')
    assert os.path.isfile(ens_path)

    pcc_df = read_ods(pcc_path, 1, headers=True)
    ens_df = read_ods(ens_path, 1, headers=True)

    feature_values = {} # {name: [values]}

    pcc_confidences = {}
    ens_confidences = {}

    pcc_values = {}
    ens_values = {}

    for row in pcc_df.to_dict(orient='records'):
        name = row['name']
        confidence = row['pcc_zero']
        best_value = row['best_value']
        feature_values.setdefault(name, [])
        feature_values[name].append(confidence)
        pcc_values[name] = best_value
        pcc_confidences[name] = confidence

    for row in ens_df.to_dict(orient='records'):
        name = row['name']
        confidence = row['confidence']
        best_value = row['best value']
        if not best_value:
            confidence = 1 - confidence
        feature_values.setdefault(name, [])
        feature_values[name].append(confidence)
        ens_values[name] = best_value
        ens_confidences[name] = confidence

    final_values = [] # [(score, name)]
    for name, confidences in feature_values.items():
        mean_confidence = sum(confidences) / len(confidences)
        final_values.append((mean_confidence, name))
    final_values.sort()
    for name, conf in final_values:
        print(name, conf)

    combine_fn = os.path.abspath(f'{current_path}/../reports/{dt}/combine-{dt}.ods')
    print(f'Saving {combine_fn}.')
    data = OrderedDict()
    rows = []
    pcc_value = 0
    ens_value = 0
    rec_change = 0
    last_value = 0
    for i, (conf, name) in enumerate(final_values, 2):
        pcc_conf = pcc_confidences.get(name, 0.5)
        ens_conf = ens_confidences.get(name, 0.5)
        algo_conf = (pcc_conf + ens_conf) / 2
        if name in last_data:
            pcc_value = pcc_values.get(name, '')
            ens_value = ens_values.get(name, '')
            last_value = last_data[name]
            should_take = bool(round(algo_conf))
        rows.append([
            name, conf, pcc_conf, ens_conf, 0.5, f'=(C{i}+D{i}+E{i})/3', algo_conf, last_value, pcc_value, ens_value, f'=IF(C{i}>D{i}, I{i}, J{i})',
            f'=IF(IF(B{i}>0.5, IF(C{i}>D{i}, I{i}, J{i})<>H{i}, H{i}<>0), TRUE, FALSE)'
        ])
    data.update({"Sheet 1": [
        ['name', 'score', 'pcc_conf', 'ens_conf', 'human_conf', 'agg_conf', 'algo_conf', 'last_value', 'pcc_value', 'ens_value', 'rec_value', 'rec_change']] + \
        rows
                 })
    save_data(combine_fn, data)
    print(f'Saved {combine_fn}.')

    auto_size_and_fix_columns(combine_fn, '-final')


if __name__ == "__main__":
    combine()
